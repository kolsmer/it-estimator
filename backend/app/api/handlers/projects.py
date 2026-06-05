from __future__ import annotations

import csv
import json
from io import BytesIO, StringIO

from fastapi import APIRouter, HTTPException
from fastapi.responses import StreamingResponse

from backend.app.internal.repository.project_storage import (
    archive_project,
    delete_project,
    get_project,
    list_active_projects,
    list_archived_projects,
    restore_project,
)


router = APIRouter()


@router.get('/projects')
def get_projects():
    return list_active_projects()


@router.get('/projects/archived')
def get_archived_projects():
    return list_archived_projects()


@router.get('/projects/{project_id}')
def get_project_detail(project_id: int):
    project = get_project(project_id)
    if not project:
        raise HTTPException(status_code=404, detail='Project not found')
    return project


@router.get('/projects/{project_id}/export')
def export_project(project_id: int, format: str = 'json'):
    project = get_project(project_id)
    if not project:
        raise HTTPException(status_code=404, detail='Project not found')

    export_format = format.lower()
    filename_base = f'project_{project_id}'

    if export_format == 'json':
        payload = json.dumps(project, ensure_ascii=False, indent=2)
        return StreamingResponse(
            StringIO(payload),
            media_type='application/json; charset=utf-8',
            headers={'Content-Disposition': f'attachment; filename="{filename_base}.json"'},
        )

    if export_format == 'csv':
        output = StringIO()
        writer = csv.DictWriter(output, fieldnames=list(project.keys()))
        writer.writeheader()
        writer.writerow(project)
        output.seek(0)
        return StreamingResponse(
            output,
            media_type='text/csv; charset=utf-8',
            headers={'Content-Disposition': f'attachment; filename="{filename_base}.csv"'},
        )

    if export_format == 'xlsx':
        try:
            from openpyxl import Workbook
        except ImportError:
            raise HTTPException(status_code=500, detail='openpyxl is not installed')

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = 'project'
        headers = list(project.keys())
        for col, header in enumerate(headers, start=1):
            sheet.cell(row=1, column=col, value=header)
        for col, header in enumerate(headers, start=1):
            sheet.cell(row=2, column=col, value=str(project.get(header, '')))

        stream = BytesIO()
        workbook.save(stream)
        stream.seek(0)
        return StreamingResponse(
            stream,
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': f'attachment; filename="{filename_base}.xlsx"'},
        )

    if export_format == 'xls':
        try:
            import xlwt
        except ImportError:
            raise HTTPException(status_code=500, detail='xlwt is not installed')

        workbook = xlwt.Workbook()
        sheet = workbook.add_sheet('project')
        headers = list(project.keys())
        for col, header in enumerate(headers):
            sheet.write(0, col, header)
        for col, header in enumerate(headers):
            sheet.write(1, col, str(project.get(header, '')))

        stream = BytesIO()
        workbook.save(stream)
        stream.seek(0)
        return StreamingResponse(
            stream,
            media_type='application/vnd.ms-excel',
            headers={'Content-Disposition': f'attachment; filename="{filename_base}.xls"'},
        )

    raise HTTPException(status_code=400, detail='Unsupported format. Use: json, csv, xls, xlsx')


@router.post('/projects/{project_id}/archive')
def archive_project_endpoint(project_id: int):
    project = archive_project(project_id)
    if not project:
        raise HTTPException(status_code=404, detail='Project not found')
    return {'success': True, 'project': project, 'message': 'Проект архивирован'}


@router.post('/projects/{project_id}/restore')
def restore_project_endpoint(project_id: int):
    project = restore_project(project_id)
    if not project:
        raise HTTPException(status_code=404, detail='Project not found in archive')
    return {'success': True, 'project': project, 'message': 'Проект восстановлен'}


@router.delete('/projects/{project_id}')
def delete_project_endpoint(project_id: int):
    deleted = delete_project(project_id)
    if not deleted:
        raise HTTPException(status_code=404, detail='Project not found')
    return {'success': True, 'message': 'Проект удалён'}
